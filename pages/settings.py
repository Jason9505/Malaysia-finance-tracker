# pages/settings.py
# Settings page — Appearance, Data Management, Backup, Reset, About.
#
# FIXES applied:
#   Code — sys.path block removed (handled by pages/__init__.py)
#   UX   — Import now tracks skipped rows and shows a detailed error report
#   Code — FONT_UI used from config instead of hardcoded "Segoe UI"

import sys
import os
import io
import shutil
import sqlite3
import zipfile
from datetime import date, datetime

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import config
from config  import (C_BG, C_CARD, C_TEXT, C_TEXT_MED, C_TEXT_LT,
                     C_SUCCESS, C_DANGER, C_WARNING, C_BORDER,
                     C_PRIMARY, C_PRIMARY_LT, C_ACCENT,
                     DB_PATH, APP_DIR, RECEIPTS_DIR, FONT_UI)
from widgets import Card, ScrollFrame, make_button
from export  import export_to_excel, export_to_zip

try:
    import openpyxl
    OPENPYXL = True
except ImportError:
    OPENPYXL = False


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _section_label(parent, text, bg):
    tk.Label(parent, text=text,
             font=(FONT_UI[0], 13, "bold"), bg=bg, fg=C_TEXT
             ).pack(anchor="w", padx=28, pady=(24, 6))


def _divider(parent, bg):
    tk.Frame(parent, bg=C_BORDER, height=1).pack(fill="x", padx=28, pady=(0, 4))


def _row_label(parent, title, subtitle="", bg=C_CARD):
    f = tk.Frame(parent, bg=bg)
    f.pack(side="left", fill="both", expand=True)
    tk.Label(f, text=title, font=(FONT_UI[0], 10, "bold"),
             bg=bg, fg=C_TEXT, anchor="w").pack(anchor="w")
    if subtitle:
        tk.Label(f, text=subtitle, font=(FONT_UI[0], 8),
                 bg=bg, fg=C_TEXT_MED, anchor="w", wraplength=420,
                 justify="left").pack(anchor="w", pady=(1, 0))


# ─────────────────────────────────────────────────────────────────────────────
# Import helpers
# ─────────────────────────────────────────────────────────────────────────────

def _import_rows(db, wb, receipt_map):
    """
    FIX (UX): Now collects skipped rows with reasons instead of swallowing them.
    Returns (inc_count, exp_count, rel_count, skipped_list).
    skipped_list is a list of (sheet, row_num, reason) tuples.
    """
    inc_count = exp_count = rel_count = 0
    skipped = []   # FIX: was silently discarded

    # ── Income sheet ──────────────────────────────────────────────────────────
    if "Income" in wb.sheetnames:
        ws   = wb["Income"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            for rn, row in enumerate(rows[1:], start=2):
                try:
                    padded = (list(row) + [""] * 7)[:7]
                    _, cat, name, amount, dt, notes, rec_base = padded
                    if not name or not amount or not dt:
                        skipped.append(("Income", rn, "Missing name, amount, or date"))
                        continue
                    if str(name).strip().upper() == "TOTAL":
                        continue
                    amount = float(str(amount).replace(",", "").strip())
                    if amount <= 0:
                        skipped.append(("Income", rn, f"Non-positive amount: {amount}"))
                        continue
                    cat    = str(cat).strip().lower() if cat else "salary"
                    dt     = str(dt).strip()[:10]
                    notes  = str(notes).strip() if notes else ""
                    rec_path = receipt_map.get(str(rec_base).strip(), "") if rec_base else ""
                    db.add_income(cat, str(name).strip(), amount, dt, notes, rec_path)
                    inc_count += 1
                except Exception as e:
                    skipped.append(("Income", rn, str(e)))

    # ── Expenses sheet ────────────────────────────────────────────────────────
    if "Expenses" in wb.sheetnames:
        from config import EXPENSE_CATS
        label_to_key = {v[0].lower(): k for k, v in EXPENSE_CATS.items()}
        ws   = wb["Expenses"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            for rn, row in enumerate(rows[1:], start=2):
                try:
                    padded = (list(row) + [""] * 8)[:8]
                    _, cat_label, name, amount, dt, tax_relief, notes, rec_base = padded
                    if not name or not amount or not dt:
                        skipped.append(("Expenses", rn, "Missing name, amount, or date"))
                        continue
                    if str(name).strip().upper() == "TOTAL":
                        continue
                    amount  = float(str(amount).replace(",", "").strip())
                    if amount <= 0:
                        skipped.append(("Expenses", rn, f"Non-positive amount: {amount}"))
                        continue
                    cat_str = str(cat_label).strip().lower() if cat_label else ""
                    cat_key = label_to_key.get(cat_str, "others")
                    dt      = str(dt).strip()[:10]
                    notes   = str(notes).strip() if notes else ""
                    tax_rel = str(tax_relief).strip() if tax_relief else ""
                    rec_path = receipt_map.get(str(rec_base).strip(), "") if rec_base else ""
                    db.add_expense(cat_key, str(name).strip(), amount, dt,
                                   notes, rec_path, tax_rel)
                    exp_count += 1
                except Exception as e:
                    skipped.append(("Expenses", rn, str(e)))

    # ── Tax Reliefs sheet — manual entries detail only ────────────────────────
    if "Tax Reliefs" in wb.sheetnames:
        ws   = wb["Tax Reliefs"]
        rows = list(ws.iter_rows(values_only=True))
        detail_start = None
        for i, row in enumerate(rows):
            first = str(row[0]).strip() if row[0] else ""
            if "manual relief entries detail" in first.lower():
                detail_start = i + 2
                break

        if detail_start is not None:
            for rn, row in enumerate(rows[detail_start:], start=detail_start + 1):
                try:
                    padded = (list(row) + [""] * 6)[:6]
                    relief_key, name, amount, dt, notes, rec_base = padded
                    if not relief_key or not amount or not dt:
                        skipped.append(("Tax Reliefs", rn, "Missing key, amount, or date"))
                        continue
                    amount = float(str(amount).replace(",", "").strip())
                    if amount <= 0:
                        skipped.append(("Tax Reliefs", rn, f"Non-positive amount: {amount}"))
                        continue
                    dt       = str(dt).strip()[:10]
                    notes    = str(notes).strip() if notes else ""
                    rec_path = receipt_map.get(str(rec_base).strip(), "") if rec_base else ""
                    db.add_relief(str(relief_key).strip(),
                                  str(name).strip() if name else str(relief_key).strip(),
                                  amount, dt, notes, rec_path)
                    rel_count += 1
                except Exception as e:
                    skipped.append(("Tax Reliefs", rn, str(e)))

    return inc_count, exp_count, rel_count, skipped


def _import_from_excel(db, filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    result = _import_rows(db, wb, receipt_map={})
    wb.close()
    return result


def _import_from_zip(db, filepath):
    rec_count   = 0
    receipt_map = {}

    with zipfile.ZipFile(filepath, "r") as zf:
        names = zf.namelist()

        os.makedirs(RECEIPTS_DIR, exist_ok=True)
        for name in names:
            if name.startswith("receipts/") and not name.endswith("/"):
                base = os.path.basename(name)
                dest = os.path.join(RECEIPTS_DIR, base)
                if not os.path.exists(dest):
                    with zf.open(name) as src, open(dest, "wb") as dst:
                        shutil.copyfileobj(src, dst)
                receipt_map[base] = dest
                rec_count += 1

        xlsx_names = [n for n in names if n.endswith(".xlsx")]
        if not xlsx_names:
            raise ValueError("No .xlsx file found inside the ZIP.")

        with zf.open(xlsx_names[0]) as xf:
            xlsx_data = xf.read()

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_data), read_only=True, data_only=True)
    inc, exp, rel, skipped = _import_rows(db, wb, receipt_map)
    wb.close()
    return inc, exp, rel, rec_count, skipped


# ─────────────────────────────────────────────────────────────────────────────
# SettingsPage
# ─────────────────────────────────────────────────────────────────────────────

class SettingsPage(ScrollFrame):

    def __init__(self, parent, db):
        super().__init__(parent)
        self.db = db
        self._theme_var = tk.StringVar(
            value=self.db.get_setting("theme", "light"))
        self._build()

    # ── Build ─────────────────────────────────────────────────────────────────

    def _build(self):
        inner = self.inner
        bg    = C_BG

        tk.Label(inner, text="Settings",
                 font=(FONT_UI[0], 20, "bold"), bg=bg, fg=C_TEXT
                 ).pack(anchor="w", padx=28, pady=(24, 0))
        tk.Label(inner, text="Manage your app preferences and data.",
                 font=(FONT_UI[0], 10), bg=bg, fg=C_TEXT_MED
                 ).pack(anchor="w", padx=28, pady=(2, 8))

        self._build_appearance(inner, bg)
        self._build_data(inner, bg)
        self._build_backup(inner, bg)
        self._build_reset(inner, bg)
        self._build_about(inner, bg)

        tk.Frame(inner, bg=bg, height=40).pack()

    # ── Section: Appearance ───────────────────────────────────────────────────

    def _build_appearance(self, parent, bg):
        _section_label(parent, "🎨  Appearance", bg)

        card = Card(parent, padx=0, pady=0,
                    highlightbackground=C_BORDER, highlightthickness=1)
        card.pack(fill="x", padx=28, pady=(0, 4))

        row = tk.Frame(card, bg=C_CARD, padx=20, pady=16)
        row.pack(fill="x")
        _row_label(row, "App Theme",
                   "Choose between Light, Dark, or follow the OS setting. "
                   "Restart the app to apply the new theme.",
                   bg=C_CARD)

        btn_frame = tk.Frame(row, bg=C_CARD)
        btn_frame.pack(side="right", padx=(16, 0))

        themes = [("☀️  Light", "light"), ("🌙  Dark", "dark"), ("🖥️  System", "system")]
        self._theme_btns = {}
        for label, val in themes:
            b = tk.Radiobutton(
                btn_frame, text=label,
                variable=self._theme_var, value=val,
                font=(FONT_UI[0], 9), bg=C_CARD, fg=C_TEXT,
                selectcolor=C_PRIMARY_LT, activebackground=C_CARD,
                cursor="hand2", command=self._on_theme_change,
                relief="flat", bd=0, indicatoron=True,
            )
            b.pack(side="left", padx=6)
            self._theme_btns[val] = b

        tk.Frame(card, bg=C_BORDER, height=1).pack(fill="x", padx=20)

        preview_row = tk.Frame(card, bg=C_CARD, padx=20, pady=12)
        preview_row.pack(fill="x")
        tk.Label(preview_row, text="Preview:",
                 font=(FONT_UI[0], 8, "bold"), bg=C_CARD, fg=C_TEXT_MED
                 ).pack(side="left", padx=(0, 10))

        self._preview_frames = {}
        preview_defs = [
            ("light",  "#f1f5f9", "#ffffff", "Light"),
            ("dark",   "#0f172a", "#1e293b", "Dark"),
            ("system", "#e2e8f0", "#f8fafc", "System"),
        ]
        for key, bg_col, card_col, lbl in preview_defs:
            pf = tk.Frame(preview_row, bg=bg_col,
                          width=72, height=40,
                          highlightthickness=2,
                          highlightbackground=C_BORDER,
                          cursor="hand2")
            pf.pack(side="left", padx=4)
            pf.pack_propagate(False)
            inner_card = tk.Frame(pf, bg=card_col, width=40, height=26)
            inner_card.place(x=4, y=7)
            inner_card.pack_propagate(False)
            for xi, dc in enumerate(["#4f46e5", "#10b981", "#ef4444"]):
                tk.Frame(inner_card, bg=dc, width=8, height=4).place(x=3 + xi * 12, y=4)
            tk.Label(pf, text=lbl, font=(FONT_UI[0], 6),
                     bg=bg_col, fg="#94a3b8").place(x=0, y=30, relwidth=1.0)
            pf.bind("<Button-1>",
                    lambda _e, v=key: (self._theme_var.set(v), self._on_theme_change()))
            inner_card.bind("<Button-1>",
                            lambda _e, v=key: (self._theme_var.set(v), self._on_theme_change()))
            self._preview_frames[key] = pf

        self._highlight_preview()

        self._restart_lbl = tk.Label(
            card,
            text="  ℹ️  Theme saved. Please restart the app to apply the new theme.",
            font=(FONT_UI[0], 9, "italic"),
            bg="#fefce8", fg="#92400e", pady=8)

    def _on_theme_change(self):
        theme = self._theme_var.get()
        self.db.set_setting("theme", theme)
        self._highlight_preview()
        self._restart_lbl.pack(fill="x", padx=20, pady=(0, 8))

    def _highlight_preview(self):
        active = self._theme_var.get()
        for key, pf in self._preview_frames.items():
            pf.config(
                highlightbackground=C_PRIMARY if key == active else C_BORDER,
                highlightthickness=2 if key == active else 1,
            )

    # ── Section: Data ─────────────────────────────────────────────────────────

    def _build_data(self, parent, bg):
        _section_label(parent, "📂  Data Management", bg)

        card = Card(parent, padx=0, pady=0,
                    highlightbackground=C_BORDER, highlightthickness=1)
        card.pack(fill="x", padx=28, pady=(0, 4))

        self._setting_row(
            card, icon="📊", title="Export to Excel  (.xlsx)",
            subtitle="Save all income, expenses, relief entries and a tax summary "
                     "to a formatted spreadsheet. Receipt files are not included.",
            btn_text="Export…", btn_bg="#f0fdf4", btn_fg=C_SUCCESS,
            command=self._do_export_excel,
        )
        tk.Frame(card, bg=C_BORDER, height=1).pack(fill="x", padx=20)
        self._setting_row(
            card, icon="📦", title="Export with Receipts  (.zip)",
            subtitle="Save everything in a single ZIP file — the full spreadsheet "
                     "plus every receipt image and PDF attached to your records. "
                     "This ZIP can be fully re-imported to restore all data and receipts.",
            btn_text="Export ZIP…", btn_bg="#f0fdf4", btn_fg=C_SUCCESS,
            command=self._do_export_zip,
            btn_state="normal" if OPENPYXL else "disabled",
        )
        tk.Frame(card, bg=C_BORDER, height=1).pack(fill="x", padx=20)

        imp_sub = ("Import from a .xlsx or .zip file previously exported by this app. "
                   "Existing data is kept — rows are appended.\n"
                   "• .zip export: receipts are fully restored.\n"
                   "• .xlsx export: data only — receipts will show as missing.")
        if not OPENPYXL:
            imp_sub += "\n⚠  openpyxl not installed. Run:  pip install openpyxl"

        self._setting_row(
            card, icon="📤", title="Import  (.xlsx or .zip)",
            subtitle=imp_sub,
            btn_text="Import…", btn_bg="#eff6ff", btn_fg=C_PRIMARY,
            command=self._do_import,
            btn_state="normal" if OPENPYXL else "disabled",
        )

    def _setting_row(self, parent, icon, title, subtitle,
                     btn_text, btn_bg, btn_fg, command, btn_state="normal"):
        row = tk.Frame(parent, bg=C_CARD, padx=20, pady=14)
        row.pack(fill="x")
        lbl_frame = tk.Frame(row, bg=C_CARD)
        lbl_frame.pack(side="left", fill="both", expand=True)
        tk.Label(lbl_frame, text=f"{icon}  {title}",
                 font=(FONT_UI[0], 10, "bold"), bg=C_CARD, fg=C_TEXT,
                 anchor="w").pack(anchor="w")
        tk.Label(lbl_frame, text=subtitle,
                 font=(FONT_UI[0], 8), bg=C_CARD, fg=C_TEXT_MED,
                 anchor="w", wraplength=500, justify="left"
                 ).pack(anchor="w", pady=(2, 0))
        make_button(row, btn_text, command,
                    bg=btn_bg, fg=btn_fg,
                    font=(FONT_UI[0], 9, "bold"), pady=6, padx=14,
                    state=btn_state,
                    ).pack(side="right", padx=(16, 0))

    def _do_export_excel(self):
        export_to_excel(self.db, self)

    def _do_export_zip(self):
        export_to_zip(self.db, self)

    def _do_import(self):
        if not OPENPYXL:
            messagebox.showerror(
                "Missing Library",
                "openpyxl is not installed.\n\nRun:  pip install openpyxl",
                parent=self)
            return

        filepath = filedialog.askopenfilename(
            title="Select file to import (.xlsx or .zip)",
            filetypes=[
                ("Supported files", "*.xlsx *.zip"),
                ("Excel Workbook",  "*.xlsx"),
                ("ZIP Archive",     "*.zip"),
                ("All Files",       "*.*"),
            ],
            parent=self)
        if not filepath:
            return

        ext    = os.path.splitext(filepath)[1].lower()
        is_zip = (ext == ".zip")

        confirm_msg = (
            "Rows from the selected file will be APPENDED to your existing data.\n\n"
            "No existing records will be deleted.\n\n"
        )
        if is_zip:
            confirm_msg += "Receipt files found in the ZIP will be copied to your receipts folder.\n\n"
        confirm_msg += "Continue?"

        if not messagebox.askyesno("Confirm Import", confirm_msg, parent=self):
            return

        try:
            if is_zip:
                inc, exp, rel, rec, skipped = _import_from_zip(self.db, filepath)
            else:
                inc, exp, rel, skipped = _import_from_excel(self.db, filepath)
                rec = 0
        except Exception as err:
            messagebox.showerror("Import Failed",
                                 f"Could not read file:\n{err}", parent=self)
            return

        if inc == 0 and exp == 0 and rel == 0:
            messagebox.showwarning(
                "Nothing Imported",
                "No valid rows were found in the file.\n\n"
                "Make sure you are importing a file exported by this app.",
                parent=self)
            return

        # FIX (UX): Show detailed summary including skipped row report
        msg = (f"Successfully imported:\n\n"
               f"  💵  Income entries:      {inc}\n"
               f"  🧾  Expense entries:     {exp}\n"
               f"  📋  Relief entries:      {rel}\n")
        if is_zip:
            msg += f"  📎  Receipt files:      {rec}\n"

        if skipped:
            msg += f"\n⚠  {len(skipped)} row(s) skipped due to errors:\n"
            for sheet, row_num, reason in skipped[:10]:  # cap at 10 lines
                msg += f"  • {sheet} row {row_num}: {reason}\n"
            if len(skipped) > 10:
                msg += f"  … and {len(skipped) - 10} more. Check your source file."

        messagebox.showinfo("Import Complete", msg, parent=self)
        self._notify_all()

    # ── Section: Backup ───────────────────────────────────────────────────────

    def _build_backup(self, parent, bg):
        _section_label(parent, "🗄️  Database Backup", bg)

        card = Card(parent, padx=0, pady=0,
                    highlightbackground=C_BORDER, highlightthickness=1)
        card.pack(fill="x", padx=28, pady=(0, 4))

        self._setting_row(
            card, icon="💾", title="Backup Database",
            subtitle="Save a copy of your tracker.db file to any folder. "
                     "Keep this safe — it contains all your financial data.",
            btn_text="Backup…", btn_bg="#f0fdf4", btn_fg=C_SUCCESS,
            command=self._do_backup,
        )
        tk.Frame(card, bg=C_BORDER, height=1).pack(fill="x", padx=20)
        self._setting_row(
            card, icon="♻️", title="Restore Database",
            subtitle="⚠️  Replace the current database with a previously saved backup. "
                     "ALL current data will be overwritten. "
                     "The app will close automatically after restore.",
            btn_text="Restore…", btn_bg="#fff1f2", btn_fg=C_DANGER,
            command=self._do_restore,
        )
        tk.Frame(card, bg=C_BORDER, height=1).pack(fill="x", padx=20)

        info_row = tk.Frame(card, bg=C_CARD, padx=20, pady=10)
        info_row.pack(fill="x")
        try:
            size_kb  = os.path.getsize(DB_PATH) / 1024
            size_str = f"{size_kb:.1f} KB"
        except OSError:
            size_str = "unknown"
        tk.Label(info_row,
                 text=f"📍 DB location: {DB_PATH}   •   Size: {size_str}",
                 font=(FONT_UI[0], 8), bg=C_CARD, fg=C_TEXT_LT,
                 anchor="w").pack(anchor="w")

    def _do_backup(self):
        today_str = date.today().strftime("%Y%m%d")
        dest = filedialog.asksaveasfilename(
            title="Save Database Backup",
            defaultextension=".db",
            filetypes=[("SQLite Database", "*.db"), ("All Files", "*.*")],
            initialfile=f"tracker_backup_{today_str}.db",
            parent=self)
        if not dest:
            return
        try:
            shutil.copy2(DB_PATH, dest)
            messagebox.showinfo("Backup Complete",
                                f"Database backed up to:\n{dest}", parent=self)
        except Exception as err:
            messagebox.showerror("Backup Failed",
                                 f"Could not copy file:\n{err}", parent=self)

    def _do_restore(self):
        src = filedialog.askopenfilename(
            title="Select Database Backup to Restore",
            filetypes=[("SQLite Database", "*.db"), ("All Files", "*.*")],
            parent=self)
        if not src:
            return
        try:
            conn = sqlite3.connect(src)
            tables = {r[0] for r in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
            conn.close()
            required = {"income", "expenses", "relief_entries", "settings"}
            if not required.issubset(tables):
                messagebox.showerror(
                    "Invalid Backup",
                    "The selected file does not appear to be a valid "
                    "Finance Tracker database backup.\n\n"
                    f"Expected tables: {', '.join(sorted(required))}",
                    parent=self)
                return
        except Exception as err:
            messagebox.showerror("Invalid File",
                                 f"Cannot read the selected file:\n{err}", parent=self)
            return

        if not messagebox.askyesno(
                "⚠️  Confirm Restore",
                "This will REPLACE your current database with the backup.\n\n"
                "ALL current data will be permanently overwritten.\n\n"
                "The application will close after the restore.\n\n"
                "Are you absolutely sure?",
                parent=self):
            return

        try:
            self.db.conn.close()
            shutil.copy2(src, DB_PATH)
            messagebox.showinfo(
                "Restore Complete",
                "Database restored successfully.\n\nThe app will now close. "
                "Please reopen it.",
                parent=self)
            self.winfo_toplevel().destroy()
        except Exception as err:
            messagebox.showerror("Restore Failed",
                                 f"Could not restore file:\n{err}", parent=self)

    # ── Section: Reset Data ───────────────────────────────────────────────────

    def _build_reset(self, parent, bg):
        _section_label(parent, "🗑️  Reset Data", bg)

        card = Card(parent, padx=0, pady=0,
                    highlightbackground=C_DANGER, highlightthickness=1)
        card.pack(fill="x", padx=28, pady=(0, 4))

        banner = tk.Frame(card, bg="#fff1f2", padx=20, pady=10)
        banner.pack(fill="x")
        tk.Label(banner,
                 text="⚠️  Danger Zone — these actions cannot be undone.",
                 font=(FONT_UI[0], 9, "bold"), bg="#fff1f2", fg=C_DANGER,
                 anchor="w").pack(anchor="w")

        tk.Frame(card, bg=C_BORDER, height=1).pack(fill="x", padx=20)

        self._setting_row(
            card, icon="🗑️", title="Reset All Data",
            subtitle="Permanently delete every income record, expense, manual relief "
                     "entry and receipt file. App settings (theme) are kept. "
                     "You will be offered a chance to export first.",
            btn_text="Reset All…", btn_bg="#fff1f2", btn_fg=C_DANGER,
            command=self._do_reset,
        )

    def _do_reset(self):
        step1 = messagebox.askyesno(
            "⚠️  Reset All Data",
            "This will permanently delete:\n\n"
            "  • All income records\n"
            "  • All expense records\n"
            "  • All manual relief entries\n"
            "  • All saved receipt files\n\n"
            "App settings (theme) will NOT be affected.\n\n"
            "Do you want to continue?",
            parent=self, icon="warning",
        )
        if not step1:
            return

        export_choice = _ExportOfferDialog(self).result
        if export_choice == "cancel":
            return
        if export_choice == "export":
            export_to_excel(self.db, self)
            if not messagebox.askyesno(
                    "Continue with Reset?",
                    "Export finished (or was skipped).\n\n"
                    "Do you still want to reset all data?",
                    parent=self, icon="warning"):
                return

        if not _TypeConfirmDialog(self, confirm_word="RESET").confirmed:
            return

        try:
            cur_settings = {"theme": self.db.get_setting("theme", "light")}
            self.db.conn.executescript("""
                DELETE FROM income;
                DELETE FROM expenses;
                DELETE FROM relief_entries;
            """)
            self.db.conn.commit()
            for k, v in cur_settings.items():
                self.db.set_setting(k, v)
        except Exception as err:
            messagebox.showerror("Reset Failed",
                                 f"Database error:\n{err}", parent=self)
            return

        deleted_files = 0
        if os.path.isdir(RECEIPTS_DIR):
            for fname in os.listdir(RECEIPTS_DIR):
                fpath = os.path.join(RECEIPTS_DIR, fname)
                try:
                    if os.path.isfile(fpath):
                        os.remove(fpath)
                        deleted_files += 1
                except OSError:
                    pass

        self._notify_all()
        messagebox.showinfo(
            "Reset Complete",
            f"All financial data has been deleted.\n\n"
            f"  🗑️  Receipt files removed: {deleted_files}",
            parent=self,
        )

    # ── Section: About ────────────────────────────────────────────────────────

    def _build_about(self, parent, bg):
        _section_label(parent, "ℹ️  About", bg)

        card = Card(parent, padx=20, pady=20,
                    highlightbackground=C_BORDER, highlightthickness=1)
        card.pack(fill="x", padx=28, pady=(0, 4))

        info_rows = [
            ("App Name",        "MyKad Financial Tracker"),
            ("Edition",         "Malaysia  —  YA 2025 tax bands"),
            ("Python",          sys.version.split()[0]),
            ("Data Folder",     APP_DIR),
            ("Receipts Folder", RECEIPTS_DIR),
            ("Database",        DB_PATH),
        ]

        grid = tk.Frame(card, bg=C_CARD)
        grid.pack(fill="x")
        grid.columnconfigure(1, weight=1)

        for i, (label, value) in enumerate(info_rows):
            tk.Label(grid, text=label,
                     font=(FONT_UI[0], 9, "bold"), bg=C_CARD, fg=C_TEXT_MED,
                     anchor="w", width=16
                     ).grid(row=i, column=0, sticky="w", pady=3)
            tk.Label(grid, text=value,
                     font=(FONT_UI[0], 9), bg=C_CARD, fg=C_TEXT,
                     anchor="w"
                     ).grid(row=i, column=1, sticky="w", pady=3, padx=(8, 0))

        tk.Frame(card, bg=C_BORDER, height=1).pack(fill="x", pady=(12, 8))

        make_button(card, "📁  Open Data Folder",
                    command=self._open_data_folder,
                    bg=C_BG, fg=C_TEXT,
                    font=(FONT_UI[0], 9), pady=5, padx=10
                    ).pack(anchor="w")

    def _open_data_folder(self):
        import platform, subprocess
        try:
            system = platform.system()
            if system == "Windows":
                os.startfile(APP_DIR)
            elif system == "Darwin":
                subprocess.Popen(["open", APP_DIR])
            else:
                subprocess.Popen(["xdg-open", APP_DIR])
        except Exception as err:
            messagebox.showerror("Error",
                                 f"Cannot open folder:\n{err}", parent=self)

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _notify_all(self):
        root = self.winfo_toplevel()
        if hasattr(root, "refresh_page"):
            for key in ("dashboard", "income", "expenses", "tax"):
                root.refresh_page(key)

    def refresh(self):
        self._theme_var.set(self.db.get_setting("theme", "light"))
        self._highlight_preview()


# ─────────────────────────────────────────────────────────────────────────────
# Helper dialogs
# ─────────────────────────────────────────────────────────────────────────────

class _ExportOfferDialog(tk.Toplevel):

    def __init__(self, parent):
        super().__init__(parent)
        self.result = "cancel"
        self.title("Export Before Reset?")
        self.configure(bg=C_CARD)
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)
        self.bind("<Escape>", lambda _e: self._choose_cancel())
        self._build()
        self._center(parent)
        self.wait_window()

    def _build(self):
        hdr = tk.Frame(self, bg=C_WARNING, padx=20, pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text="📤  Export Your Data First?",
                 font=(FONT_UI[0], 13, "bold"), bg=C_WARNING, fg="white"
                 ).pack(side="left")

        body = tk.Frame(self, bg=C_CARD, padx=28, pady=20)
        body.pack(fill="both")

        tk.Label(body,
                 text="Before deleting everything, would you like to\n"
                      "export your data to an Excel file for safekeeping?",
                 font=(FONT_UI[0], 10), bg=C_CARD, fg=C_TEXT,
                 justify="center").pack(pady=(0, 6))
        tk.Label(body,
                 text="You can re-import this file later using\n"
                      "Settings → Import from Excel.",
                 font=(FONT_UI[0], 9, "italic"), bg=C_CARD, fg=C_TEXT_MED,
                 justify="center").pack(pady=(0, 20))

        make_button(body, "📥  Yes — Export to Excel first",
                    command=self._choose_export,
                    bg="#f0fdf4", fg=C_SUCCESS,
                    font=(FONT_UI[0], 10, "bold"), pady=10,
                    ).pack(fill="x", pady=(0, 6))
        make_button(body, "⏭️  No — Skip export and continue",
                    command=self._choose_skip,
                    bg="#fff7ed", fg=C_WARNING,
                    font=(FONT_UI[0], 10), pady=8,
                    ).pack(fill="x", pady=(0, 6))
        make_button(body, "✕  Cancel — Don't reset anything",
                    command=self._choose_cancel,
                    bg=C_BG, fg=C_TEXT_MED,
                    font=(FONT_UI[0], 10), pady=8,
                    ).pack(fill="x")

    def _choose_export(self): self.result = "export"; self.destroy()
    def _choose_skip(self):   self.result = "skip";   self.destroy()
    def _choose_cancel(self): self.result = "cancel"; self.destroy()

    def _center(self, parent):
        self.update_idletasks()
        pw = parent.winfo_width();  ph = parent.winfo_height()
        px = parent.winfo_rootx(); py = parent.winfo_rooty()
        dw = self.winfo_width();   dh = self.winfo_height()
        self.geometry(f"+{px + (pw - dw) // 2}+{py + (ph - dh) // 2}")


class _TypeConfirmDialog(tk.Toplevel):

    def __init__(self, parent, confirm_word="RESET"):
        super().__init__(parent)
        self.confirmed = False
        self._word     = confirm_word
        self.title("Final Confirmation")
        self.configure(bg=C_CARD)
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)
        self.bind("<Escape>", lambda _e: self.destroy())
        self._build()
        self._center(parent)
        self.wait_window()

    def _build(self):
        hdr = tk.Frame(self, bg=C_DANGER, padx=20, pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🗑️  Final Confirmation",
                 font=(FONT_UI[0], 13, "bold"), bg=C_DANGER, fg="white"
                 ).pack(side="left")

        body = tk.Frame(self, bg=C_CARD, padx=28, pady=22)
        body.pack(fill="both")

        tk.Label(body,
                 text=f"Type  {self._word}  below to confirm deletion.",
                 font=(FONT_UI[0], 10), bg=C_CARD, fg=C_TEXT
                 ).pack(anchor="w", pady=(0, 6))
        tk.Label(body,
                 text="This action is PERMANENT and cannot be undone.",
                 font=(FONT_UI[0], 9, "bold"), bg=C_CARD, fg=C_DANGER
                 ).pack(anchor="w", pady=(0, 14))

        self._entry_var = tk.StringVar()
        entry = tk.Entry(body, textvariable=self._entry_var,
                         font=(FONT_UI[0], 13, "bold"),
                         relief="solid", bd=2, bg="white",
                         justify="center",
                         highlightthickness=2, highlightcolor=C_DANGER, fg=C_DANGER)
        entry.pack(fill="x", ipady=8, pady=(0, 16))
        entry.focus_set()
        self._entry_var.trace_add("write", self._on_type)

        self._confirm_btn = make_button(
            body, "🗑️  Delete All Data", command=self._confirm,
            bg=C_DANGER, fg="white",
            font=(FONT_UI[0], 11, "bold"), pady=10, state="disabled")
        self._confirm_btn.pack(fill="x", pady=(0, 6))

        make_button(body, "Cancel", command=self.destroy,
                    bg=C_BG, fg=C_TEXT_MED,
                    font=(FONT_UI[0], 10), pady=6).pack(fill="x")

    def _on_type(self, *_):
        typed = self._entry_var.get()
        self._confirm_btn.config(state="normal" if typed == self._word else "disabled")

    def _confirm(self):
        self.confirmed = True
        self.destroy()

    def _center(self, parent):
        self.update_idletasks()
        pw = parent.winfo_width();  ph = parent.winfo_height()
        px = parent.winfo_rootx(); py = parent.winfo_rooty()
        dw = self.winfo_width();   dh = self.winfo_height()
        self.geometry(f"+{px + (pw - dw) // 2}+{py + (ph - dh) // 2}")