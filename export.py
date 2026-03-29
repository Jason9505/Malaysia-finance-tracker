# export.py
# Excel export: Income, Expenses, Tax Reliefs, Tax Summary sheets.
# ZIP export:   same Excel + all receipt files bundled together.
# Requires:  python -m pip install openpyxl

import os
import io
import zipfile
from datetime import date
from tkinter import filedialog, messagebox

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL = True
except ImportError:
    OPENPYXL = False

from config import ALL_RELIEFS, EXPENSE_CATS
from utils  import compute_full_tax


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_col):
    return PatternFill("solid", fgColor=hex_col.lstrip("#"))

def _font(bold=False, colour="000000", size=10):
    return Font(bold=bold, color=colour.lstrip("#"), size=size)

def _center():
    return Alignment(horizontal="center", vertical="center")

def _header(ws, labels, row=1, bg="4f46e5", fg="FFFFFF"):
    for c, label in enumerate(labels, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font      = _font(bold=True, colour=fg)
        cell.fill      = _fill(bg)
        cell.alignment = _center()

def _widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


# ── Sheet builders ────────────────────────────────────────────────────────────
# When with_receipts=True every data row gets an extra "Receipt File" column
# containing the basename of the receipt (e.g. receipt_20250101_120000.png).
# The ZIP exporter then bundles those files under receipts/<basename>.

def _income_sheet(wb, db, with_receipts=False):
    ws = wb.create_sheet("Income")
    headers = ["#", "Category", "Name", "Amount (RM)", "Date", "Notes"]
    widths  = {"A":5,"B":14,"C":32,"D":16,"E":14,"F":36}
    if with_receipts:
        headers.append("Receipt File")
        widths["G"] = 36
    _header(ws, headers, bg="10b981")
    _widths(ws, widths)

    rows = db.get_income()
    for i, r in enumerate(rows, 1):
        rec_base = os.path.basename(r["receipt"]) if r["receipt"] else ""
        row_data = [i, r["category"].capitalize(), r["name"],
                    r["amount"], r["date"], r["notes"] or ""]
        if with_receipts:
            row_data.append(rec_base)
        ws.append(row_data)
        ws.cell(row=i+1, column=4).number_format = "#,##0.00"

    if rows:
        tr = len(rows) + 2
        ws.cell(row=tr, column=3, value="TOTAL").font = _font(bold=True)
        tc = ws.cell(row=tr, column=4, value=db.total_income())
        tc.font = _font(bold=True, colour="10b981")
        tc.number_format = "#,##0.00"


def _expenses_sheet(wb, db, with_receipts=False):
    ws = wb.create_sheet("Expenses")
    headers = ["#", "Category", "Name", "Amount (RM)", "Date", "Tax Relief", "Notes"]
    widths  = {"A":5,"B":22,"C":32,"D":16,"E":14,"F":20,"G":36}
    if with_receipts:
        headers.append("Receipt File")
        widths["H"] = 36
    _header(ws, headers, bg="ef4444")
    _widths(ws, widths)

    rows = db.get_expenses()
    for i, r in enumerate(rows, 1):
        cat      = EXPENSE_CATS.get(r["category"], (r["category"],))[0]
        rec_base = os.path.basename(r["receipt"]) if r["receipt"] else ""
        row_data = [i, cat, r["name"], r["amount"],
                    r["date"], r["tax_relief"] or "", r["notes"] or ""]
        if with_receipts:
            row_data.append(rec_base)
        ws.append(row_data)
        ws.cell(row=i+1, column=4).number_format = "#,##0.00"

    if rows:
        tr = len(rows) + 2
        ws.cell(row=tr, column=3, value="TOTAL").font = _font(bold=True)
        tc = ws.cell(row=tr, column=4, value=db.total_expenses())
        tc.font = _font(bold=True, colour="ef4444")
        tc.number_format = "#,##0.00"


def _reliefs_sheet(wb, db, with_receipts=False):
    ws = wb.create_sheet("Tax Reliefs")
    _header(ws, ["Relief Category","Max (RM)","Auto (RM)","Manual (RM)","Claimed (RM)","Notes"])
    _widths(ws, {"A":40,"B":14,"C":14,"D":14,"E":14,"F":40})

    auto_map   = db.tax_deductible_by_relief()
    manual_all = db.get_reliefs()
    manual_by_key = {}
    for r in manual_all:
        manual_by_key.setdefault(r["relief_key"], []).append(r)

    total_relief = 0.0
    for key, name, max_rm in ALL_RELIEFS:
        auto_amt   = auto_map.get(key, 0.0)
        manual_amt = sum(r["amount"] for r in manual_by_key.get(key, []))
        claimed    = min(auto_amt + manual_amt, max_rm)
        notes_list = [r["notes"] for r in manual_by_key.get(key, []) if r["notes"]]
        notes_str  = " | ".join(notes_list) if notes_list else ""
        total_relief += claimed
        ws.append([name, max_rm, auto_amt, manual_amt, claimed, notes_str])
        for col in [2, 3, 4, 5]:
            ws.cell(row=ws.max_row, column=col).number_format = "#,##0.00"

    # Manual entries detail block
    ws.append([])
    ws.append(["Manual Relief Entries Detail"])
    ws.cell(row=ws.max_row, column=1).font = _font(bold=True, colour="4f46e5")
    detail_headers = ["Relief Key","Description","Amount (RM)","Date","Notes"]
    detail_widths  = {"A":28,"B":32,"C":16,"D":14,"E":40}
    if with_receipts:
        detail_headers.append("Receipt File")
        detail_widths["F"] = 36
    _header(ws, detail_headers, row=ws.max_row+1, bg="e0e7ff", fg="1e293b")

    for r in manual_all:
        rec_base = os.path.basename(r["receipt"]) if r["receipt"] else ""
        row_data = [r["relief_key"], r["name"], r["amount"],
                    r["date"], r["notes"] or ""]
        if with_receipts:
            row_data.append(rec_base)
        ws.append(row_data)
        ws.cell(row=ws.max_row, column=3).number_format = "#,##0.00"

    return total_relief


def _tax_summary_sheet(wb, db, total_relief):
    ws = wb.create_sheet("Tax Summary")
    _widths(ws, {"A":36,"B":20})

    gross  = db.total_income("salary")
    result = compute_full_tax(gross, total_relief)

    data = [
        ("Malaysia Income Tax Summary (YA 2025)", None),
        (f"Generated: {date.today().strftime('%d %B %Y')}", None),
        ("", None),
        ("Gross Taxable Income (Salary)",  result["gross"]),
        ("Total Tax Reliefs",              result["reliefs"]),
        ("Chargeable Income",              result["chargeable"]),
        ("", None),
        ("Income Tax (Before Rebate)",     result["tax_before"]),
        ("Tax Rebate",                     result["rebate"]),
        ("Net Tax Payable",                result["net_tax"]),
        ("Effective Tax Rate",             f"{result['eff_rate']:.2f}%"),
    ]
    for label, value in data:
        ws.append([label, value])

    ws["A1"].font = _font(bold=True, size=13, colour="4f46e5")
    ws["A2"].font = _font(colour="64748b")
    for ri, colour in [(4,"000000"),(5,"10b981"),(6,"000000"),
                       (8,"f59e0b"),(9,"10b981"),(10,"ef4444")]:
        ws.cell(row=ri, column=1).font = _font(bold=True)
        c = ws.cell(row=ri, column=2)
        if isinstance(c.value, float):
            c.number_format = "#,##0.00"
            c.font = _font(bold=True, colour=colour)


# ── Collect all receipt paths referenced in the DB ───────────────────────────

def _collect_receipt_paths(db):
    """
    Return a dict  basename → absolute_path  for every receipt file that
    exists on disk.  Duplicates (same basename) keep the first path found.
    """
    receipts = {}
    all_rows = list(db.get_income()) + list(db.get_expenses()) + list(db.get_reliefs())
    for row in all_rows:
        try:
            path = row["receipt"] or ""
        except (IndexError, KeyError):
            path = ""
        if path and os.path.isfile(path):
            base = os.path.basename(path)
            if base not in receipts:
                receipts[base] = path
    return receipts


# ── Public API ────────────────────────────────────────────────────────────────

def export_to_excel(db, parent_window=None):
    """Prompt user for a save path then write a plain .xlsx file (no receipts)."""
    if not OPENPYXL:
        messagebox.showerror(
            "Missing Library",
            "openpyxl is not installed.\n\n"
            "Run this in your terminal:\n\n"
            "    python -m pip install openpyxl",
            parent=parent_window)
        return

    today_str = date.today().strftime("%Y%m%d")
    filepath  = filedialog.asksaveasfilename(
        title="Export to Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook","*.xlsx"), ("All Files","*.*")],
        initialfile=f"finance_tracker_{today_str}.xlsx",
        parent=parent_window)
    if not filepath:
        return

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _income_sheet(wb, db, with_receipts=False)
        _expenses_sheet(wb, db, with_receipts=False)
        total_relief = _reliefs_sheet(wb, db, with_receipts=False)
        _tax_summary_sheet(wb, db, total_relief)
        wb.save(filepath)
        messagebox.showinfo("Export Complete",
                            f"File saved to:\n{filepath}",
                            parent=parent_window)
    except Exception as exc:
        messagebox.showerror("Export Failed",
                             f"Could not save file:\n{exc}",
                             parent=parent_window)


def export_to_zip(db, parent_window=None):
    """
    Export a .zip file containing:
      • finance_tracker.xlsx  — full spreadsheet with a 'Receipt File' column
      • receipts/<filename>   — every receipt image/PDF that exists on disk

    The ZIP can be re-imported by this app to fully restore all data + receipts.
    """
    if not OPENPYXL:
        messagebox.showerror(
            "Missing Library",
            "openpyxl is not installed.\n\n"
            "Run this in your terminal:\n\n"
            "    python -m pip install openpyxl",
            parent=parent_window)
        return

    today_str = date.today().strftime("%Y%m%d")
    filepath  = filedialog.asksaveasfilename(
        title="Export with Receipts (ZIP)",
        defaultextension=".zip",
        filetypes=[("ZIP Archive","*.zip"), ("All Files","*.*")],
        initialfile=f"finance_tracker_{today_str}.zip",
        parent=parent_window)
    if not filepath:
        return

    try:
        # Build Excel workbook in memory
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _income_sheet(wb, db, with_receipts=True)
        _expenses_sheet(wb, db, with_receipts=True)
        total_relief = _reliefs_sheet(wb, db, with_receipts=True)
        _tax_summary_sheet(wb, db, total_relief)

        xlsx_buf = io.BytesIO()
        wb.save(xlsx_buf)
        xlsx_buf.seek(0)

        # Gather receipt files
        receipt_map = _collect_receipt_paths(db)

        # Write ZIP
        missing = []
        with zipfile.ZipFile(filepath, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("finance_tracker.xlsx", xlsx_buf.read())
            for base, abs_path in receipt_map.items():
                try:
                    zf.write(abs_path, f"receipts/{base}")
                except Exception:
                    missing.append(base)

        total_rec = len(receipt_map)
        added_rec = total_rec - len(missing)

        msg = f"ZIP saved to:\n{filepath}\n\n"
        msg += f"  📊  Excel spreadsheet: included\n"
        msg += f"  📎  Receipt files:      {added_rec} of {total_rec} included"
        if missing:
            msg += f"\n  ⚠   {len(missing)} file(s) could not be read and were skipped."
        messagebox.showinfo("Export Complete", msg, parent=parent_window)

    except Exception as exc:
        messagebox.showerror("Export Failed",
                             f"Could not create ZIP:\n{exc}",
                             parent=parent_window)